"""
Excel exporter for POC-3C — Application Review Mode.

Produces ``application_review.xlsx`` with four sheets:

    - Review Summary      — one row per job; key verdict columns
    - Filled Fields       — every field that was filled (or would be filled)
    - Unknown Fields      — every field with no bank answer
    - Required Missing    — required fields that blocked submission
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from loguru import logger

from app.models.application_review import ApplicationReviewRecord


class ApplicationReviewExporter:
    """Export a list of ApplicationReviewRecord objects to an Excel workbook."""

    def __init__(self, export_dir: Path) -> None:
        self._export_dir = export_dir

    def export(self, records: list[ApplicationReviewRecord]) -> Path:
        """Write the four-sheet workbook and return its path."""
        self._export_dir.mkdir(parents=True, exist_ok=True)
        filepath = self._export_dir / "application_review.xlsx"

        # ── Sheet 1: Review Summary ─────────────────────────────────────────
        summary_rows = []
        for r in records:
            mode = "DRY_RUN" if r.dry_run else "LIVE"
            summary_rows.append({
                "Job ID":               r.job_id,
                "Job Title":            r.job_title,
                "Company":              r.company,
                "Mode":                 mode,
                "Total Fields":         r.total_fields,
                "Filled Fields":        r.filled_count,
                "Unknown Fields":       r.unknown_count,
                "Required Missing":     r.missing_required_count,
                "Fill Rate (%)":        r.fill_rate_pct,
                "Ready To Submit":      r.ready_to_submit_label,
                "Screenshot (Final)":   r.screenshot_final_state or "",
                "Reviewed At":          r.reviewed_at.strftime("%Y-%m-%d %H:%M:%S"),
            })

        # ── Sheet 2: Filled Fields ──────────────────────────────────────────
        filled_rows = []
        for r in records:
            mode = "DRY_RUN" if r.dry_run else "LIVE"
            for f in r.filled_fields:
                filled_rows.append({
                    "Job ID":       r.job_id,
                    "Company":      r.company,
                    "Job Title":    r.job_title,
                    "Mode":         mode,
                    "Question Key": f.question_key,
                    "Question Text":f.question_text,
                    "Field Type":   f.field_type,
                    "Required":     "Yes" if f.required else "No",
                    "Status":       f.status,
                    "Value Used":   f.answer_used or "",
                })

        # ── Sheet 3: Unknown Fields ─────────────────────────────────────────
        unknown_rows = []
        for r in records:
            mode = "DRY_RUN" if r.dry_run else "LIVE"
            for u in r.unknown_fields:
                unknown_rows.append({
                    "Job ID":       r.job_id,
                    "Company":      r.company,
                    "Job Title":    r.job_title,
                    "Mode":         mode,
                    "Question Key": u.question_key,
                    "Question Text":u.question_text,
                    "Field Type":   u.field_type,
                    "Required":     "Yes" if u.required else "No",
                })

        # ── Sheet 4: Required Missing ───────────────────────────────────────
        missing_rows = []
        for r in records:
            for m in r.required_fields_missing:
                missing_rows.append({
                    "Job ID":          r.job_id,
                    "Company":         r.company,
                    "Job Title":       r.job_title,
                    "Ready To Submit": r.ready_to_submit_label,
                    "Question Key":    m.question_key,
                    "Question Text":   m.question_text,
                    "Field Type":      m.field_type,
                    "Action Required": "Add answer to candidate_profile.json or answers.py",
                })

        # ── DataFrames ──────────────────────────────────────────────────────
        _cols_summary = [
            "Job ID", "Job Title", "Company", "Mode",
            "Total Fields", "Filled Fields", "Unknown Fields",
            "Required Missing", "Fill Rate (%)", "Ready To Submit",
            "Screenshot (Final)", "Reviewed At",
        ]
        _cols_filled  = [
            "Job ID", "Company", "Job Title", "Mode",
            "Question Key", "Question Text", "Field Type",
            "Required", "Status", "Value Used",
        ]
        _cols_unknown = [
            "Job ID", "Company", "Job Title", "Mode",
            "Question Key", "Question Text", "Field Type", "Required",
        ]
        _cols_missing = [
            "Job ID", "Company", "Job Title", "Ready To Submit",
            "Question Key", "Question Text", "Field Type", "Action Required",
        ]

        df_summary = pd.DataFrame(summary_rows) if summary_rows else pd.DataFrame(columns=_cols_summary)
        df_filled  = pd.DataFrame(filled_rows)  if filled_rows  else pd.DataFrame(columns=_cols_filled)
        df_unknown = pd.DataFrame(unknown_rows) if unknown_rows else pd.DataFrame(columns=_cols_unknown)
        df_missing = pd.DataFrame(missing_rows) if missing_rows else pd.DataFrame(columns=_cols_missing)

        with pd.ExcelWriter(str(filepath), engine="openpyxl") as writer:
            df_summary.to_excel(writer, index=False, sheet_name="Review Summary")
            df_filled.to_excel(writer,  index=False, sheet_name="Filled Fields")
            df_unknown.to_excel(writer, index=False, sheet_name="Unknown Fields")
            df_missing.to_excel(writer, index=False, sheet_name="Required Missing")

            for sheet_name, df in [
                ("Review Summary",  df_summary),
                ("Filled Fields",   df_filled),
                ("Unknown Fields",  df_unknown),
                ("Required Missing",df_missing),
            ]:
                ws = writer.sheets[sheet_name]
                for col_idx, col in enumerate(df.columns, start=1):
                    max_len = max(
                        len(str(col)),
                        df[col].astype(str).str.len().max() if not df.empty else 0,
                    )
                    ws.column_dimensions[
                        ws.cell(row=1, column=col_idx).column_letter
                    ].width = min(max(max_len + 2, 14), 80)

            # Highlight Ready To Submit column (green=YES, red=NO)
            from openpyxl.styles import PatternFill
            green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            ws_summary = writer.sheets["Review Summary"]
            ready_col = list(df_summary.columns).index("Ready To Submit") + 1
            for row_idx in range(2, len(df_summary) + 2):
                cell = ws_summary.cell(row=row_idx, column=ready_col)
                cell.fill = green if cell.value == "YES" else red

        ready_count = sum(1 for r in records if r.ready_to_submit)
        logger.info(
            "Application review exported: {} ({}/{} ready to submit)",
            filepath.name,
            ready_count,
            len(records),
        )
        return filepath
